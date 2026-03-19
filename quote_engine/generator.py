"""
MSP Quote Generator — core PDF engine.
Public API: generate_pdf(file_obj) -> BytesIO
"""

from io import BytesIO
from datetime import datetime
import openpyxl
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable, KeepTogether, PageBreak,
)
from reportlab.pdfgen import canvas

# ── CONSTANTS ─────────────────────────────────────────────────────────────────
DETAIL_SHEET  = "MSP LMS"
SUMMARY_SHEET = "Client Summary (2)"

COMPANY_NAME    = "MSP Construction Ltd"
COMPANY_ADDRESS = "Brighton, East Sussex"
COMPANY_EMAIL   = "info@mspconstruction.co.uk"
COMPANY_PHONE   = "+44 (0)1273 000 000"
COMPANY_REG     = "Company Reg: 12345678  |  VAT Reg: GB 123 456 789"

FOOTER_TEXT = (
    "Quotation valid 30 days from date stated. All works subject to MSP Construction Ltd "
    "standard terms and conditions. Prices exclude VAT. Variations require written instruction."
)

PAGE_SIZE = landscape(A4)
MARGIN    = 14 * mm

C_GREEN     = colors.HexColor("#2c4a3e")
C_GOLD      = colors.HexColor("#b8965a")
C_CREAM     = colors.HexColor("#faf8f3")
C_CREAM_ALT = colors.HexColor("#f2efe6")
C_DARK      = colors.HexColor("#1a1a1a")
C_MID       = colors.HexColor("#555555")
C_RULE      = colors.HexColor("#d8d0c0")
C_WHITE     = colors.white

F_HEAD  = "Times-Bold"
F_SERIF = "Times-Roman"
F_BODY  = "Helvetica"
F_BOLD  = "Helvetica-Bold"
F_MONO  = "Courier"
F_MONOB = "Courier-Bold"


# ── DATA CLASSES ──────────────────────────────────────────────────────────────
class SummaryItem:
    def __init__(self, num, desc, price):
        self.num, self.desc, self.price = num, desc, price

class SubItem:
    def __init__(self, num, desc, note, labour, material, subcon,
                 activity_cost, prelims, prelims_pct, po, sell):
        self.num           = num
        self.desc          = desc
        self.note          = note
        self.labour        = labour
        self.material      = material
        self.subcon        = subcon
        self.activity_cost = activity_cost
        self.prelims       = prelims
        self.prelims_pct   = prelims_pct
        self.po            = po
        self.sell          = sell

class Section:
    def __init__(self, num, title):
        self.num, self.title = num, title
        self.items = []
        self.tot_labour = self.tot_material = self.tot_subcon = 0.0
        self.tot_act    = self.tot_prelims  = self.tot_prelims_pct = 0.0
        self.tot_po     = self.tot_sell     = 0.0


# ── HELPERS ───────────────────────────────────────────────────────────────────
def sf(v):
    try:    return float(v)
    except: return 0.0

def fmt(v, zero="—"):
    if v is None or v == 0: return zero
    return f"£{v:,.2f}"

def fmtp(v):
    if not v: return "—"
    return f"{float(v)*100:.2f}%"

def is_sec(s):
    if not isinstance(s, str): return False
    p = s.strip().split(".")
    return len(p) == 2 and p[1] == "0" and p[0].isdigit()

def is_sub(s):
    if not isinstance(s, str): return False
    p = s.strip().split(".")
    return len(p) == 2 and p[1] != "0" and p[0].isdigit()

def fmt_date(dt):
    return dt.strftime("%d %B %Y").lstrip("0")


# ── READ EXCEL ────────────────────────────────────────────────────────────────
def read_excel(file_obj):
    """Accept a path string or any file-like object (BytesIO, SpooledTemporaryFile, etc.)."""
    wb = openpyxl.load_workbook(file_obj, data_only=True)

    # ── Client Summary (2) ────────────────────────────────────────────────────
    ws1 = wb[SUMMARY_SHEET]
    hdr = {"date": None, "client": None, "address": None, "notes": None, "start_date": "TBC"}
    summary_items = []

    for row in ws1.iter_rows(values_only=True):
        if not any(c is not None for c in row):
            continue
        a = str(row[0]).strip() if row[0] else ""
        b = row[1]
        d = row[3] if len(row) > 3 else None

        if a == "Date" and b:
            hdr["date"] = fmt_date(b) if isinstance(b, datetime) else str(b)
        elif a == "Client":
            hdr["client"] = str(b).strip() if b else ""
        elif a == "Site Address":
            hdr["address"] = str(b).strip() if b else ""
        elif a == "Notes on Property":
            hdr["notes"] = str(b).strip() if b else ""
        elif a == "Start Date":
            hdr["start_date"] = str(b).strip() if b else "TBC"
        elif is_sec(a) or a == "0.0":
            price = float(d) if isinstance(d, (int, float)) else None
            summary_items.append(SummaryItem(a, str(b).strip() if b else "", price))
        elif a == "" and isinstance(b, str) and "TOTAL" in b.upper():
            summary_items.append(SummaryItem("", "TOTAL EXCL VAT", sf(d)))

    if "0.0" not in [si.num for si in summary_items]:
        summary_items.insert(0, SummaryItem("0.0", "Preliminaries", None))

    grand_total = next((si.price for si in summary_items if si.desc == "TOTAL EXCL VAT"), 0.0)

    # ── MSP LMS ───────────────────────────────────────────────────────────────
    ws2   = wb[DETAIL_SHEET]
    rows2 = list(ws2.iter_rows(values_only=True))

    detail_start = None
    for i, row in enumerate(rows2):
        r = list(row) + [None] * 35
        if (isinstance(r[1], str) and r[1].strip() == "1.0"
                and isinstance(r[2], str) and "Prelim" in r[2]
                and isinstance(r[28], str)):
            detail_start = i
            break

    sections, cur = [], None

    if detail_start is not None:
        for row in rows2[detail_start:]:
            r    = list(row) + [None] * 35
            b, c, d = r[1], r[2], r[3]
            if not isinstance(b, str): continue
            item = b.strip()
            desc = str(c).strip() if c else ""

            if is_sec(item):
                if cur is not None: sections.append(cur)
                cur = Section(num=item, title=desc)
                continue

            if cur is None: continue

            if desc == "Activity Total":
                cur.tot_labour      = sf(r[7])
                cur.tot_material    = sf(r[13])
                cur.tot_subcon      = sf(r[18])
                cur.tot_act         = sf(r[20])
                cur.tot_prelims     = sf(r[22])
                cur.tot_prelims_pct = sf(r[24])
                cur.tot_po          = sf(r[26])
                cur.tot_sell        = sf(r[28])
                continue

            if desc in ("Description", "TOTAL EXCL VAT", "TOTAL FOR ALL WORKS"): continue
            if r[5] is not None and desc.lower().startswith("labour"): continue

            if is_sub(item):
                cur.items.append(SubItem(
                    num=item, desc=desc,
                    note=str(d).strip() if d else None,
                    labour=sf(r[7]), material=sf(r[13]), subcon=sf(r[18]),
                    activity_cost=sf(r[20]), prelims=sf(r[22]),
                    prelims_pct=sf(r[24]), po=sf(r[26]), sell=sf(r[28]),
                ))

        if cur is not None: sections.append(cur)

    for sec in sections:
        if sec.num == "1.0" and sec.tot_sell == 0.0:
            sec.tot_sell = sum(it.sell for it in sec.items)

    return hdr, summary_items, grand_total, sections


# ── PAGE CHROME ───────────────────────────────────────────────────────────────
class QuoteCanvas(canvas.Canvas):
    def __init__(self, *args, ci=None, **kw):
        super().__init__(*args, **kw)
        self._ci = ci or {}

    def showPage(self): self._chrome(); super().showPage()
    def save(self):     self._chrome(); super().save()

    def _chrome(self):
        w, h = PAGE_SIZE
        self.setFillColor(C_GREEN)
        self.rect(0, h - 20*mm, w, 20*mm, fill=1, stroke=0)
        self.setFillColor(C_GOLD)
        self.rect(0, h - 20*mm - 1*mm, w, 1*mm, fill=1, stroke=0)
        self.setFillColor(C_WHITE)
        self.setFont(F_HEAD, 15)
        self.drawString(MARGIN, h - 13*mm, self._ci.get("name", COMPANY_NAME))
        self.setFont(F_BODY, 7)
        self.setFillColor(colors.HexColor("#aaccbb"))
        self.drawRightString(w - MARGIN, h - 8*mm,
            self._ci.get("phone", "") + "   " + self._ci.get("email", ""))
        self.drawRightString(w - MARGIN, h - 14*mm, self._ci.get("address", ""))
        self.setFillColor(C_GOLD)
        self.rect(MARGIN, 14*mm, w - MARGIN*2, 0.35*mm, fill=1, stroke=0)
        self.setFont(F_BODY, 6)
        self.setFillColor(C_MID)
        self.drawString(MARGIN, 10*mm, FOOTER_TEXT)
        self.drawRightString(w - MARGIN, 10*mm, COMPANY_REG)
        self.setFont(F_MONO, 6)
        self.setFillColor(colors.HexColor("#aaaaaa"))
        self.drawCentredString(w / 2, 6*mm, f"Page {self._pageNumber}")


# ── STYLES ────────────────────────────────────────────────────────────────────
def PS(name, **kw): return ParagraphStyle(name, **kw)

def make_styles():
    return {
        "title":   PS("t_title",  fontName=F_HEAD,  fontSize=20, textColor=C_GREEN, leading=24, spaceAfter=2),
        "lbl":     PS("t_lbl",    fontName=F_BOLD,  fontSize=7.5, textColor=C_MID,  leading=10, spaceAfter=1),
        "val":     PS("t_val",    fontName=F_BODY,  fontSize=10, textColor=C_DARK,  leading=13, spaceAfter=2),
        "val_lg":  PS("t_val_lg", fontName=F_SERIF, fontSize=12, textColor=C_DARK,  leading=16),
        "scope":   PS("t_scope",  fontName=F_BODY,  fontSize=9.5, textColor=C_DARK, leading=14, spaceAfter=4),
        "note":    PS("t_note",   fontName=F_BODY,  fontSize=7.5, textColor=C_MID,  leading=11, spaceAfter=3),
        "s_hl":    PS("t_s_hl",   fontName=F_BOLD,  fontSize=8,  textColor=C_WHITE),
        "s_hr":    PS("t_s_hr",   fontName=F_BOLD,  fontSize=8,  textColor=C_WHITE, alignment=TA_RIGHT),
        "s_num":   PS("t_s_num",  fontName=F_MONO,  fontSize=9,  textColor=C_MID,   alignment=TA_CENTER),
        "s_desc":  PS("t_s_desc", fontName=F_BODY,  fontSize=9.5,textColor=C_DARK),
        "s_prc":   PS("t_s_prc",  fontName=F_MONO,  fontSize=9.5,textColor=C_DARK,  alignment=TA_RIGHT),
        "s_tl":    PS("t_s_tl",   fontName=F_HEAD,  fontSize=11, textColor=C_WHITE, alignment=TA_RIGHT),
        "s_tr":    PS("t_s_tr",   fontName=F_MONOB, fontSize=11, textColor=C_WHITE, alignment=TA_RIGHT),
        "d_hl":    PS("t_d_hl",   fontName=F_BOLD,  fontSize=7,  textColor=C_WHITE),
        "d_hr":    PS("t_d_hr",   fontName=F_BOLD,  fontSize=7,  textColor=C_WHITE, alignment=TA_RIGHT),
        "d_hc":    PS("t_d_hc",   fontName=F_BOLD,  fontSize=7,  textColor=C_WHITE, alignment=TA_CENTER),
        "sec_n":   PS("t_sec_n",  fontName=F_MONOB, fontSize=9,  textColor=C_WHITE),
        "sec_t":   PS("t_sec_t",  fontName=F_HEAD,  fontSize=10, textColor=C_WHITE),
        "sec_v":   PS("t_sec_v",  fontName=F_MONOB, fontSize=9,  textColor=C_WHITE, alignment=TA_RIGHT),
        "d_num":   PS("t_d_num",  fontName=F_MONO,  fontSize=7.5,textColor=C_MID,   alignment=TA_CENTER),
        "d_dsc":   PS("t_d_dsc",  fontName=F_BODY,  fontSize=8,  textColor=C_DARK,  leading=10),
        "d_val":   PS("t_d_val",  fontName=F_MONO,  fontSize=7.5,textColor=C_DARK,  alignment=TA_RIGHT),
        "d_pct":   PS("t_d_pct",  fontName=F_MONO,  fontSize=7.5,textColor=C_MID,   alignment=TA_RIGHT),
        "d_inc":   PS("t_d_inc",  fontName=F_BODY,  fontSize=7,  textColor=C_MID,   alignment=TA_RIGHT),
        "at_l":    PS("t_at_l",   fontName=F_BOLD,  fontSize=8,  textColor=C_WHITE, alignment=TA_RIGHT),
        "at_v":    PS("t_at_v",   fontName=F_MONOB, fontSize=8,  textColor=C_WHITE, alignment=TA_RIGHT),
        "at_p":    PS("t_at_p",   fontName=F_MONOB, fontSize=8,  textColor=C_WHITE, alignment=TA_RIGHT),
    }


def det_cols(uw):
    return [12*mm, 68*mm, 22*mm, 22*mm, 23*mm, 23*mm, 22*mm, 16*mm, 22*mm, 24*mm]


# ── BUILD PDF → BytesIO ───────────────────────────────────────────────────────
def build_pdf(hdr, summary_items, grand_total, sections) -> BytesIO:
    buf = BytesIO()
    w, h = PAGE_SIZE
    doc = SimpleDocTemplate(
        buf, pagesize=PAGE_SIZE,
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=27*mm, bottomMargin=22*mm,
        title=f"Quote – {hdr.get('client', '')}",
        author=COMPANY_NAME,
    )
    ST    = make_styles()
    uw    = w - MARGIN * 2
    story = []

    # ── PART 1: Client Summary ────────────────────────────────────────────────
    story.append(Paragraph("QUOTATION", ST["title"]))
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_GOLD, spaceAfter=5))

    ref_date = hdr.get("date") or fmt_date(datetime.today())
    ref      = _make_ref(ref_date)

    left_t = Table([
        [Paragraph("CLIENT",                      ST["lbl"])],
        [Paragraph(hdr.get("client", "—"),        ST["val_lg"])],
        [Spacer(1, 4)],
        [Paragraph("SITE ADDRESS",                ST["lbl"])],
        [Paragraph(hdr.get("address", "—"),       ST["val"])],
        [Spacer(1, 4)],
        [Paragraph("ANTICIPATED START",           ST["lbl"])],
        [Paragraph(hdr.get("start_date", "TBC"),  ST["val"])],
    ], colWidths=[uw * 0.38])

    right_t = Table([
        [Paragraph("QUOTE REF",   ST["lbl"]),  Paragraph("DATE",   ST["lbl"])],
        [Paragraph(ref,           ST["val"]),   Paragraph(ref_date, ST["val"])],
        [Spacer(1, 8), ""],
        [Paragraph("PREPARED BY", ST["lbl"]),  Paragraph("", ST["lbl"])],
        [Paragraph(COMPANY_NAME,  ST["val"]),  Paragraph("", ST["val"])],
        [Paragraph(COMPANY_EMAIL, PS("rce", fontName=F_BODY, fontSize=8.5, textColor=C_MID)), ""],
        [Paragraph(COMPANY_PHONE, PS("rcp", fontName=F_BODY, fontSize=8.5, textColor=C_MID)), ""],
    ], colWidths=[uw * 0.22, uw * 0.18])

    scope_t = Table([
        [Paragraph("SCOPE OF WORKS", PS("sow", fontName=F_HEAD, fontSize=9,
                                         textColor=C_GREEN, spaceAfter=2))],
        [HRFlowable(width="100%", thickness=0.3, color=C_RULE, spaceAfter=3)],
        [Paragraph(hdr.get("notes", ""), ST["scope"])],
        [Paragraph("This quotation covers all labour, materials, plant and equipment "
                   "inclusive of preliminaries, overheads and profit.", ST["note"])],
    ], colWidths=[uw * 0.38])

    header_block = Table([[left_t, right_t, scope_t]],
                         colWidths=[uw * 0.42, uw * 0.22, uw * 0.36])
    header_block.setStyle(TableStyle([
        ("VALIGN",        (0,0),(-1,-1), "TOP"),
        ("BACKGROUND",    (0,0),(-1,-1), C_CREAM),
        ("BOX",           (0,0),(-1,-1), 0.4, C_RULE),
        ("LEFTPADDING",   (0,0),(-1,-1), 8),
        ("RIGHTPADDING",  (0,0),(-1,-1), 8),
        ("TOPPADDING",    (0,0),(-1,-1), 9),
        ("BOTTOMPADDING", (0,0),(-1,-1), 9),
        ("LINEBEFORE",    (1,0),(1,0),   0.3, C_RULE),
        ("LINEBEFORE",    (2,0),(2,0),   0.3, C_RULE),
    ]))
    story.append(header_block)
    story.append(Spacer(1, 5*mm))

    story.append(Paragraph("CLIENT SUMMARY", PS("csh", fontName=F_HEAD, fontSize=10,
                            textColor=C_GREEN, spaceAfter=2)))
    story.append(HRFlowable(width="100%", thickness=0.3, color=C_RULE, spaceAfter=4))

    cw_sum  = [14*mm, uw - 14*mm - 34*mm, 34*mm]
    sum_rows = [[
        Paragraph("Item",            ST["s_hl"]),
        Paragraph("Description",     ST["s_hl"]),
        Paragraph("Total Excl. VAT", ST["s_hr"]),
    ]]
    data_n = 0
    for si in summary_items:
        if si.desc == "TOTAL EXCL VAT": continue
        sum_rows.append([
            Paragraph(si.num,  ST["s_num"]),
            Paragraph(si.desc, ST["s_desc"]),
            Paragraph(fmt(si.price), ST["s_prc"]),
        ])
        data_n += 1
    sum_rows.append([
        Paragraph("",               ST["s_tl"]),
        Paragraph("TOTAL EXCL. VAT", ST["s_tl"]),
        Paragraph(fmt(grand_total),  ST["s_tr"]),
    ])

    sum_t = Table(sum_rows, colWidths=cw_sum, repeatRows=1)
    sts = [
        ("BACKGROUND",    (0,0),(-1,0),   C_GREEN),
        ("TOPPADDING",    (0,0),(-1,0),   7), ("BOTTOMPADDING",(0,0),(-1,0),7),
        ("TOPPADDING",    (0,1),(-1,-1),  4), ("BOTTOMPADDING",(0,1),(-1,-1),4),
        ("LEFTPADDING",   (0,0),(-1,-1),  5), ("RIGHTPADDING", (0,0),(-1,-1),5),
        ("VALIGN",        (0,0),(-1,-1),  "MIDDLE"),
        ("LINEBELOW",     (0,0),(-1,-2),  0.25, C_RULE),
        ("ALIGN",         (2,0),(2,-1),   "RIGHT"),
        ("BACKGROUND",    (0,-1),(-1,-1), C_GREEN),
        ("LINEABOVE",     (0,-1),(-1,-1), 1.5,  C_GOLD),
        ("TOPPADDING",    (0,-1),(-1,-1), 8),   ("BOTTOMPADDING",(0,-1),(-1,-1),8),
    ]
    for i in range(1, data_n + 1):
        sts.append(("BACKGROUND", (0,i),(-1,i), C_CREAM_ALT if i%2==0 else C_CREAM))
    sum_t.setStyle(TableStyle(sts))
    story.append(sum_t)

    # ── PART 2: Labour & Materials Estimate ──────────────────────────────────
    story.append(PageBreak())
    story.append(Paragraph("LABOUR &amp; MATERIALS ESTIMATE", ST["title"]))
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_GOLD, spaceAfter=5))

    cw = det_cols(uw)

    col_hdr = [
        Paragraph("Item",                    ST["d_hc"]),
        Paragraph("Description",             ST["d_hl"]),
        Paragraph("Labour\nTotal £",         ST["d_hr"]),
        Paragraph("Material\nTotal £",       ST["d_hr"]),
        Paragraph("Subcontract\nTotal £",    ST["d_hr"]),
        Paragraph("Activity\nCost",          ST["d_hr"]),
        Paragraph("Prelims",                 ST["d_hr"]),
        Paragraph("Prelims\n%",              ST["d_hr"]),
        Paragraph("Profit &amp;\nOverheads", ST["d_hr"]),
        Paragraph("Activity\nTotal",         ST["d_hr"]),
    ]

    all_rows = [col_hdr]

    for sec in sections:
        all_rows.append([
            Paragraph(sec.num,              ST["sec_n"]),
            Paragraph(sec.title.upper(),    ST["sec_t"]),
            Paragraph("", ST["sec_v"]), Paragraph("", ST["sec_v"]),
            Paragraph("", ST["sec_v"]), Paragraph("", ST["sec_v"]),
            Paragraph("", ST["sec_v"]), Paragraph("", ST["sec_v"]),
            Paragraph("", ST["sec_v"]), Paragraph("", ST["sec_v"]),
        ])

        visible = [it for it in sec.items if it.desc and it.desc not in (
            "Labour\nTotal £", "Material\nTotal £", "Subcontract\nTotal £",
            "Activity Selling price", "Activity Total",
        )]

        for it in visible:
            def _p(v): return Paragraph(fmt(v), ST["d_val"])
            price_cell = (Paragraph("Included", ST["d_inc"])
                          if it.note and it.note.lower() == "inc"
                          else _p(it.sell))
            all_rows.append([
                Paragraph(it.num,  ST["d_num"]),
                Paragraph(it.desc, ST["d_dsc"]),
                _p(it.labour), _p(it.material), _p(it.subcon),
                _p(it.activity_cost), _p(it.prelims),
                Paragraph(fmtp(it.prelims_pct), ST["d_pct"]),
                _p(it.po), price_cell,
            ])

        if visible or sec.tot_sell:
            all_rows.append([
                Paragraph("",               ST["at_l"]),
                Paragraph("Activity Total", ST["at_l"]),
                Paragraph(fmt(sec.tot_labour,      "—"), ST["at_v"]),
                Paragraph(fmt(sec.tot_material,    "—"), ST["at_v"]),
                Paragraph(fmt(sec.tot_subcon,      "—"), ST["at_v"]),
                Paragraph(fmt(sec.tot_act,         "—"), ST["at_v"]),
                Paragraph(fmt(sec.tot_prelims,     "—"), ST["at_v"]),
                Paragraph(fmtp(sec.tot_prelims_pct),     ST["at_p"]),
                Paragraph(fmt(sec.tot_po,          "—"), ST["at_v"]),
                Paragraph(fmt(sec.tot_sell,        "—"), ST["at_v"]),
            ])

    big_table = Table(all_rows, colWidths=cw, repeatRows=1)

    ts = [
        ("BACKGROUND",    (0,0),(-1,0), C_GREEN),
        ("TOPPADDING",    (0,0),(-1,0), 5), ("BOTTOMPADDING",(0,0),(-1,0),5),
        ("LEFTPADDING",   (0,0),(-1,0), 3), ("RIGHTPADDING", (0,0),(-1,0),3),
        ("VALIGN",        (0,0),(-1,0), "MIDDLE"),
        ("ALIGN",         (2,0),(-1,0), "RIGHT"),
        ("LINEBELOW",     (0,0),(-1,0), 1.0, C_GOLD),
        ("TOPPADDING",    (0,1),(-1,-1), 3), ("BOTTOMPADDING",(0,1),(-1,-1),3),
        ("LEFTPADDING",   (0,1),(-1,-1), 3), ("RIGHTPADDING", (0,1),(-1,-1),3),
        ("VALIGN",        (0,1),(-1,-1), "TOP"),
        ("ALIGN",         (2,1),(-1,-1), "RIGHT"),
    ]

    row_idx, sec_rows, act_rows, item_rows = 1, [], [], []
    for sec in sections:
        sec_rows.append(row_idx); row_idx += 1
        visible = [it for it in sec.items if it.desc and it.desc not in (
            "Labour\nTotal £", "Material\nTotal £", "Subcontract\nTotal £",
            "Activity Selling price", "Activity Total",
        )]
        for _ in visible:
            item_rows.append(row_idx); row_idx += 1
        if visible or sec.tot_sell:
            act_rows.append(row_idx); row_idx += 1

    for r in sec_rows:
        ts += [
            ("BACKGROUND",    (0,r),(-1,r), C_GREEN),
            ("LINEBELOW",     (0,r),(-1,r), 1.0, C_GOLD),
            ("TOPPADDING",    (0,r),(-1,r), 6), ("BOTTOMPADDING",(0,r),(-1,r),6),
            ("SPAN",          (1,r),(-1,r)),
        ]
    for r in act_rows:
        ts += [
            ("BACKGROUND",    (0,r),(-1,r), C_GREEN),
            ("LINEABOVE",     (0,r),(-1,r), 0.5, C_GOLD),
            ("TOPPADDING",    (0,r),(-1,r), 5), ("BOTTOMPADDING",(0,r),(-1,r),5),
        ]
    for i, r in enumerate(item_rows):
        ts += [
            ("BACKGROUND", (0,r),(-1,r), C_CREAM_ALT if i%2==0 else C_CREAM),
            ("LINEBELOW",  (0,r),(-1,r), 0.2, C_RULE),
        ]

    big_table.setStyle(TableStyle(ts))
    story.append(big_table)

    story.append(Spacer(1, 4*mm))
    tw = sum(det_cols(uw)[-4:])
    gt_t = Table([[
        Paragraph("",                ST["s_tl"]),
        Paragraph("TOTAL EXCL. VAT", ST["s_tl"]),
        Paragraph(fmt(grand_total),  ST["s_tr"]),
    ]], colWidths=[uw - tw, tw * 0.45, tw * 0.55])
    gt_t.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0), C_GREEN),
        ("LINEABOVE",     (0,0),(-1,0), 2, C_GOLD),
        ("TOPPADDING",    (0,0),(-1,0), 9), ("BOTTOMPADDING",(0,0),(-1,0),9),
        ("LEFTPADDING",   (0,0),(-1,0), 6), ("RIGHTPADDING", (0,0),(-1,0),6),
        ("ALIGN",         (1,0),(2,0),  "RIGHT"),
    ]))
    story.append(gt_t)

    ci = {"name": COMPANY_NAME, "address": COMPANY_ADDRESS,
          "email": COMPANY_EMAIL, "phone": COMPANY_PHONE}

    def make_canvas(filename, **kw):
        return QuoteCanvas(filename, pagesize=PAGE_SIZE, ci=ci)

    doc.build(story, canvasmaker=make_canvas)
    buf.seek(0)
    return buf


# ── PUBLIC API ────────────────────────────────────────────────────────────────
def generate_pdf(file_obj) -> BytesIO:
    """Main entry point. Accepts any file-like object or path string."""
    hdr, summary_items, grand_total, sections = read_excel(file_obj)
    return build_pdf(hdr, summary_items, grand_total, sections)


def _make_ref(date_str):
    try:
        dt = datetime.strptime(date_str, "%d %B %Y")
        return f"MSP-{dt.strftime('%Y%m')}-001"
    except Exception:
        return "MSP-001"


# ── CLI ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys, os
    path = sys.argv[1] if len(sys.argv) > 1 else None
    if not path or not os.path.exists(path):
        print("Usage: python generator.py workbook.xlsx"); sys.exit(1)
    with open(path, "rb") as f:
        buf = generate_pdf(f)
    out = path.replace(".xlsx", ".pdf")
    with open(out, "wb") as f:
        f.write(buf.read())
    print(f"Done: {out}")
