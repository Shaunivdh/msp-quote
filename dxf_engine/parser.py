"""
DXF → Excel parser.

Extracts labelled dimensions from DXF DIMENSION entities, spatially associates
them with nearby room/area label text, and groups by orientation (H × V) so
you get output like "Room A = 4200 mm × 3100 mm".

Also extracts:
  - Steel section callouts (UC, UB, PFC, …)
  - Window/door type labels (W-A … D-G)
  - Floor level annotations (FFL, +NNNN, GF FFL, DATUM)
"""

from __future__ import annotations

import math
import os
import re
import tempfile
from collections import defaultdict
from io import BytesIO

import ezdxf
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Palette ───────────────────────────────────────────────────────────────────
_GREEN  = "2C4A3E"
_GOLD   = "B8965A"
_ALT    = "F2EFE6"
_WHITE  = "FFFFFF"
_DARK   = "1A1A1A"
_FG     = "FFFFFF"

# ── Section classification (from layer name) ──────────────────────────────────
_SECTION_KEYWORDS: list[tuple[list[str], str]] = [
    (["ground", "gf", "g/f", "g.f"], "Ground Floor"),
    (["first", "1st", "ff", "f/f", "f.f", "level 1", "lvl1", "level1"], "First Floor"),
    (["second", "2nd", "sf", "s/f", "level 2", "lvl2", "level2"], "Second Floor"),
    (["window", "door", "schedule", "sched"], "Window/Door Schedule"),
    (["elev", "elevation", "facade", "front", "rear", "side"], "Elevations"),
    # External / overall must come before generic "General" but after specific floors
    (["external", "ext-", "_ext", "ext_", "overall", "boundary", "setback", "site plan"], "External"),
]

SECTION_ORDER = [
    "Ground Floor",
    "First Floor",
    "Second Floor",
    "Window/Door Schedule",
    "Elevations",
    "External",
    "General",
]


def _classify_layer(layer_name: str) -> str:
    name = layer_name.lower()
    for keywords, section in _SECTION_KEYWORDS:
        for kw in keywords:
            if kw in name:
                return section
    return "General"


# DXF $INSUNITS → mm scale factor
_UNITS_SCALE: dict[int, float] = {
    0: 1.0,    # unitless — assume mm
    1: 25.4,   # inches
    2: 304.8,  # feet
    4: 1.0,    # mm
    5: 10.0,   # cm
    6: 1000.0, # m
}

# ── Patterns for non-dimension text extraction ────────────────────────────────
_STEEL_RE = re.compile(
    r"(\d+\s*[xX×]\s*\d+(?:\s*[xX×]\s*\d+(?:\.\d+)?)?)\s*"
    r"(UC|UB|PFC|CHS|SHS|RHS|EA|UA|RSJ)\b",
    re.IGNORECASE,
)
_WINDOW_PATTERNS = [
    re.compile(r"\b(?:WINDOW|WIN)\s*[-–]\s*([A-G])\b", re.IGNORECASE),
    re.compile(r"\bW\s*[-–]\s*([A-G])\b", re.IGNORECASE),
]
_DOOR_PATTERNS = [
    re.compile(r"\b(?:DOOR|DR)\s*[-–]\s*([A-G])\b", re.IGNORECASE),
    re.compile(r"\bD\s*[-–]\s*([A-G])\b", re.IGNORECASE),
]
_TYPE_PATTERN = re.compile(r"\bTYPE\s*[-–:]?\s*([A-G])\b", re.IGNORECASE)
_FLOOR_RE = re.compile(
    r"(\+\s*\d{3,5}(?:\.\d+)?|GF\s*FFL|FFL\b[^\n]*|\bDATUM\b|\bFFE\b)",
    re.IGNORECASE,
)

# Text strings that are drawing admin / metadata — never a spatial label
_NON_LABEL_RE = re.compile(
    r"^(scale|nts|n\.t\.s\.|note|notes|rev\b|revision|drawing no|dwg no|date|chk|"
    r"do not scale|copyright|true north|north point|sheet \d|issued|status|"
    r"approved|checked|drawn|client|project)\b",
    re.IGNORECASE,
)


def _is_room_label(text: str) -> bool:
    """
    Return True if this text could be a spatial label (room name, area tag,
    external dim callout, etc.) worth associating with nearby dimensions.
    """
    text = text.strip()
    # Pure numeric / symbol strings are dimension values, not labels
    if re.match(r"^[\d\s.,+%@#\-/\\()\[\]\"\'m°:=x×]+$", text):
        return False
    # Too short (single letters are probably reference tags, keep them)
    # or too long (probably a paragraph note)
    if len(text) == 0 or len(text) > 80:
        return False
    if _NON_LABEL_RE.match(text):
        return False
    return True


# ── DXF helpers ───────────────────────────────────────────────────────────────

def _get_text(entity) -> str:
    try:
        if entity.dxftype() == "MTEXT":
            return entity.plain_mtext()
        return entity.dxf.text
    except Exception:
        return ""


def _iter_model_entities(doc):
    yield from doc.modelspace()
    for layout in doc.layouts:
        if layout.name.lower() != "model":
            yield from layout


def _load_doc(file_bytes: bytes):
    with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name
    try:
        return ezdxf.readfile(tmp_path)
    finally:
        os.unlink(tmp_path)


def _get_scale(doc) -> float:
    try:
        return _UNITS_SCALE.get(doc.header.get("$INSUNITS", 4), 1.0)
    except Exception:
        return 1.0


def _dim_measurement(entity) -> float | None:
    """Actual measurement: stored value first, then geometric fallback."""
    raw = entity.dxf.get("actual_measurement", 0) or 0
    if raw > 0:
        return float(raw)
    p2 = entity.dxf.get("defpoint2")
    p3 = entity.dxf.get("defpoint3")
    if p2 is None or p3 is None:
        return None
    dist = math.sqrt(sum((b - a) ** 2 for a, b in zip(p2, p3)))
    return dist if dist > 0 else None


def _dim_orientation(entity) -> str:
    """Classify a DIMENSION as Horizontal, Vertical, Diagonal, Angular, Radius, or Diameter."""
    dimtype = (entity.dxf.get("dimtype", 0) or 0) & 0x0F
    if dimtype == 2:
        return "Angular"
    if dimtype == 3:
        return "Diameter"
    if dimtype == 4:
        return "Radius"

    # Linear (0 = rotated, 1 = aligned)
    if dimtype == 0:
        angle = float(entity.dxf.get("angle", 0) or 0) % 180
    else:
        p2 = entity.dxf.get("defpoint2")
        p3 = entity.dxf.get("defpoint3")
        if p2 and p3:
            dx = float(p3[0]) - float(p2[0])
            dy = float(p3[1]) - float(p2[1])
            angle = abs(math.degrees(math.atan2(dy, dx))) % 180
        else:
            return "Linear"

    if angle < 20 or angle > 160:
        return "Horizontal"
    if 70 < angle < 110:
        return "Vertical"
    return "Diagonal"


def _dim_display_text(entity, value_mm: float) -> str:
    """Return the text string as shown on the drawing."""
    override = entity.dxf.get("text", "<>") or "<>"
    formatted = f"{value_mm:.0f}"
    if override.strip() in ("<>", ""):
        return formatted
    # "<>" inside the override is a placeholder for the measured value
    return override.replace("<>", formatted).strip()


def _dim_position(entity) -> tuple[float, float] | None:
    """Return (x, y) of the dimension text midpoint."""
    tm = entity.dxf.get("text_midpoint")
    if tm:
        return float(tm[0]), float(tm[1])
    p2 = entity.dxf.get("defpoint2")
    p3 = entity.dxf.get("defpoint3")
    if p2 and p3:
        return (float(p2[0]) + float(p3[0])) / 2, (float(p2[1]) + float(p3[1])) / 2
    return None


# ── Data collection ───────────────────────────────────────────────────────────

def _collect_text_entities(doc, source: str) -> list[dict]:
    texts = []
    for entity in _iter_model_entities(doc):
        if entity.dxftype() not in ("TEXT", "MTEXT"):
            continue
        try:
            text = _get_text(entity).strip()
            if not text:
                continue
            pos = entity.dxf.get("insert")
            texts.append(
                {
                    "text": text,
                    "x": float(pos[0]) if pos else None,
                    "y": float(pos[1]) if pos else None,
                    "layer": entity.dxf.get("layer", "0"),
                    "source": source,
                    "is_label": _is_room_label(text),
                }
            )
        except Exception:
            pass
    return texts


def _nearest_label(
    cx: float,
    cy: float,
    labels: list[dict],
    dim_value_mm: float,
    dim_section: str,
) -> str | None:
    """
    Return the text of the nearest positional label within reach of this dimension.

    Search radius = 1.5× the dimension length, floored at 1500 mm and capped at
    8000 mm.

    Section isolation:
      - External dims only match labels from External-classified layers, so they
        can't steal a room name that happens to be the geometrically nearest text.
      - Interior dims (Ground Floor, First Floor, etc.) exclude External labels.
    """
    radius = min(max(dim_value_mm * 1.5, 1500), 8000)

    is_external_dim = dim_section == "External"

    best, best_dist = None, float("inf")
    for lb in labels:
        if lb["x"] is None:
            continue
        lb_section = _classify_layer(lb["layer"])
        lb_is_external = lb_section == "External"

        # Cross-section mismatch: skip
        if is_external_dim and not lb_is_external:
            continue
        if not is_external_dim and lb_is_external:
            continue

        dist = math.sqrt((lb["x"] - cx) ** 2 + (lb["y"] - cy) ** 2)
        if dist < best_dist and dist <= radius:
            best_dist = dist
            best = lb["text"]
    return best


def _collect_dimensions(doc, scale: float, source: str, text_entities: list[dict]) -> list[dict]:
    labels = [t for t in text_entities if t["is_label"] and t["x"] is not None]
    dims = []
    for entity in _iter_model_entities(doc):
        if entity.dxftype() != "DIMENSION":
            continue
        try:
            raw = _dim_measurement(entity)
            if raw is None or raw <= 0:
                continue
            value_mm = round(raw * scale, 1)
            display = _dim_display_text(entity, value_mm)
            orientation = _dim_orientation(entity)
            pos = _dim_position(entity)
            layer = entity.dxf.get("layer", "0")

            section = _classify_layer(layer)
            nearest = None
            if pos and labels:
                nearest = _nearest_label(pos[0], pos[1], labels, value_mm, section)

            # If no label found nearby, fall back to the layer name so the
            # dimension still gets a meaningful group header in the output
            # (e.g. "EXT-DIMS" is more informative than "(unlabelled)")
            label_for_group = nearest or f"[{layer}]"

            dims.append(
                {
                    "value_mm": value_mm,
                    "display": display,
                    "orientation": orientation,
                    "x": pos[0] if pos else None,
                    "y": pos[1] if pos else None,
                    "nearest_label": nearest,
                    "label_for_group": label_for_group,
                    "layer": layer,
                    "section": section,
                    "source": source,
                }
            )
        except Exception:
            pass
    return dims


# ── Accessory data ────────────────────────────────────────────────────────────

def _extract_steel(text_entities: list[dict]) -> list[dict]:
    results, seen = [], set()
    for te in text_entities:
        for m in _STEEL_RE.finditer(te["text"]):
            size = m.group(1).replace(" ", "")
            stype = m.group(2).upper()
            label = f"{size}{stype}"
            if label not in seen:
                seen.add(label)
                results.append(
                    {"label": label, "size": size, "type": stype,
                     "layer": te["layer"], "source": te["source"],
                     "full_text": te["text"].strip()[:80]}
                )
    results.sort(key=lambda x: (x["type"], x["label"]))
    return results


def _extract_windows_doors(text_entities: list[dict]) -> list[dict]:
    results, seen = [], set()
    for te in text_entities:
        text = te["text"]
        for pat in _WINDOW_PATTERNS:
            for m in pat.finditer(text):
                letter = m.group(1).upper()
                key = ("Window", letter)
                if key not in seen:
                    seen.add(key)
                    results.append({"label": f"W-{letter}", "type": letter,
                                    "category": "Window", "layer": te["layer"],
                                    "source": te["source"], "full_text": text.strip()[:80]})
        for pat in _DOOR_PATTERNS:
            for m in pat.finditer(text):
                letter = m.group(1).upper()
                key = ("Door", letter)
                if key not in seen:
                    seen.add(key)
                    results.append({"label": f"D-{letter}", "type": letter,
                                    "category": "Door", "layer": te["layer"],
                                    "source": te["source"], "full_text": text.strip()[:80]})
        for m in _TYPE_PATTERN.finditer(text):
            letter = m.group(1).upper()
            if ("Window", letter) not in seen and ("Door", letter) not in seen:
                key = ("Unknown", letter)
                if key not in seen:
                    seen.add(key)
                    results.append({"label": f"Type {letter}", "type": letter,
                                    "category": "Unknown", "layer": te["layer"],
                                    "source": te["source"], "full_text": text.strip()[:80]})
    results.sort(key=lambda x: (x["category"], x["type"]))
    return results


def _extract_floor_heights(text_entities: list[dict]) -> list[dict]:
    results, seen = [], set()
    for te in text_entities:
        for m in _FLOOR_RE.finditer(te["text"]):
            annotation = m.group(0).strip()
            num_m = re.search(r"\+\s*(\d+(?:\.\d+)?)", annotation)
            level_mm = float(num_m.group(1)) if num_m else None
            key = annotation.upper()[:30]
            if key not in seen:
                seen.add(key)
                results.append({"annotation": annotation, "level_mm": level_mm,
                                 "layer": te["layer"], "source": te["source"],
                                 "description": te["text"].strip()[:80]})
    results.sort(key=lambda x: (x["level_mm"] is None, x["level_mm"] or 0))
    return results


# ── Excel styles ──────────────────────────────────────────────────────────────

def _hfont():  return Font(name="Calibri", bold=True, color=_FG, size=10)
def _bfont():  return Font(name="Calibri", bold=True, color=_DARK, size=10)
def _nfont():  return Font(name="Calibri", color=_DARK, size=10)
def _hfill():  return PatternFill("solid", fgColor=_GREEN)
def _sfill():  return PatternFill("solid", fgColor=_GOLD)
def _afill(i): return PatternFill("solid", fgColor=_ALT if i % 2 == 0 else _WHITE)
def _center(): return Alignment(horizontal="center", vertical="center")
def _left():   return Alignment(horizontal="left", vertical="center")

def _border():
    s = Side(style="thin", color="D8D0C0")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_row(ws, headers, widths):
    bd = _border()
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = _hfont(), _hfill(), _center(), bd
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20


def _data_row(ws, row_idx, values_aligns, i):
    bd = _border()
    fill = _afill(i)
    for col, (val, align) in enumerate(values_aligns, 1):
        c = ws.cell(row=row_idx, column=col, value=val)
        c.font, c.fill, c.alignment, c.border = _nfont(), fill, align, bd


# ── Sheet builders ────────────────────────────────────────────────────────────

def _mm_to_m(v: float) -> float:
    """Convert millimetres to metres, rounded to 3 decimal places."""
    return round(v / 1000, 3)


def _sheet_by_room(wb, dims: list[dict]):
    """
    Sheet 1: Dimensions by Room / Label.

    Each row = one label group (room name or layer fallback).
    Horizontal measurements → Width columns, Vertical → Height columns.
    Area (m²) and Perimeter (m) are computed from the primary H × V dims.

    Labels that came from spatial text association are shown as-is.
    Labels that fell back to the layer name are shown as [LAYER-NAME]
    so they're visually distinct.
    """
    ws = wb.create_sheet("Dimensions by Room")
    ws.freeze_panes = "A2"

    # Group by (label_for_group, section, layer, source)
    groups: dict = defaultdict(lambda: {"H": [], "V": [], "other": []})

    for d in dims:
        key = (d["label_for_group"], d["section"], d["layer"], d["source"])
        orient = d["orientation"]
        if orient == "Horizontal":
            groups[key]["H"].append(d["value_mm"])
        elif orient == "Vertical":
            groups[key]["V"].append(d["value_mm"])
        else:
            groups[key]["other"].append((d["value_mm"], orient))

    def _sort_key(item):
        (label, section, layer, source), _ = item
        sec_idx = SECTION_ORDER.index(section) if section in SECTION_ORDER else 99
        # Layer-fallback labels (start with "[") sort after real labels
        is_fallback = label.startswith("[")
        return (sec_idx, is_fallback, label)

    rows = []
    for (label, section, layer, source), bucket in sorted(groups.items(), key=_sort_key):
        rows.append((
            label, section, layer, source,
            sorted(bucket["H"], reverse=True),
            sorted(bucket["V"], reverse=True),
            bucket["other"],
        ))

    max_h = max((len(r[4]) for r in rows), default=1)
    max_v = max((len(r[5]) for r in rows), default=1)

    headers = (
        ["Room / Area", "Section", "Area (m²)", "Perimeter (m)", "Layer", "Source File"]
        + (["Width (m)"] + [f"H-dim {i+1} (m)" for i in range(1, max_h)])
        + (["Height (m)"] + [f"V-dim {i+1} (m)" for i in range(1, max_v)])
        + ["Other dims (m)"]
    )
    widths = [30, 20, 12, 14, 26, 22] + [12] * max_h + [12] * max_v + [32]
    _header_row(ws, headers, widths)

    bd = _border()
    for i, (label, section, layer, source, h_vals, v_vals, other) in enumerate(rows):
        row_idx = i + 2
        fill = _afill(i)

        # Compute area and perimeter from largest H and V (in metres)
        primary_h_m = _mm_to_m(h_vals[0]) if h_vals else None
        primary_v_m = _mm_to_m(v_vals[0]) if v_vals else None
        area_m2 = round(primary_h_m * primary_v_m, 3) if primary_h_m and primary_v_m else ""
        perimeter_m = round(2 * (primary_h_m + primary_v_m), 3) if primary_h_m and primary_v_m else ""

        h_vals_m = [_mm_to_m(v) for v in h_vals]
        v_vals_m = [_mm_to_m(v) for v in v_vals]

        cells = [label, section, area_m2, perimeter_m, layer, source]
        cells += h_vals_m + [""] * (max_h - len(h_vals_m))
        cells += v_vals_m + [""] * (max_v - len(v_vals_m))
        cells.append("  |  ".join(f"{_mm_to_m(v)} ({o})" for v, o in other) if other else "")

        for col, val in enumerate(cells, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.font, c.fill, c.border = _nfont(), fill, bd
            c.alignment = _center() if col > 2 else _left()

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _sheet_all_dims(wb, dims: list[dict]):
    """Sheet 2: Flat list of every dimension with full context."""
    ws = wb.create_sheet("All Dimensions")
    ws.freeze_panes = "A2"
    headers = ["#", "Display Text", "Value (m)", "Orientation",
               "Nearest Label", "Section", "Layer", "Source File"]
    widths  = [5, 20, 12, 12, 28, 20, 26, 22]
    _header_row(ws, headers, widths)

    sorted_dims = sorted(dims, key=lambda x: (
        SECTION_ORDER.index(x["section"]) if x["section"] in SECTION_ORDER else 99,
        x["nearest_label"] or "",
        -x["value_mm"],
    ))
    for i, d in enumerate(sorted_dims, 1):
        _data_row(ws, i + 1, [
            (i,                        _center()),
            (d["display"],             _left()),
            (_mm_to_m(d["value_mm"]),  _center()),
            (d["orientation"],         _center()),
            (d["label_for_group"],     _left()),
            (d["section"],             _left()),
            (d["layer"],               _left()),
            (d["source"],              _left()),
        ], i)

    ws.auto_filter.ref = f"A1:H1"


def _sheet_steel(wb, steel: list[dict]):
    ws = wb.create_sheet("Steel Sections")
    ws.freeze_panes = "A2"
    _header_row(ws, ["Label", "Size", "Type", "Layer", "Source File", "Source Text"],
                    [22, 18, 10, 26, 22, 44])
    for i, s in enumerate(steel, 1):
        _data_row(ws, i + 1, [
            (s["label"],     _left()),
            (s["size"],      _left()),
            (s["type"],      _center()),
            (s["layer"],     _left()),
            (s["source"],    _left()),
            (s["full_text"], _left()),
        ], i)


def _sheet_windows_doors(wb, items: list[dict]):
    ws = wb.create_sheet("Windows & Doors")
    ws.freeze_panes = "A2"
    _header_row(ws, ["Label", "Type", "Category", "Layer", "Source File", "Source Text"],
                    [12, 10, 12, 26, 22, 44])
    for i, wd in enumerate(items, 1):
        _data_row(ws, i + 1, [
            (wd["label"],     _left()),
            (wd["type"],      _center()),
            (wd["category"],  _left()),
            (wd["layer"],     _left()),
            (wd["source"],    _left()),
            (wd["full_text"], _left()),
        ], i)


def _sheet_floor_heights(wb, items: list[dict]):
    ws = wb.create_sheet("Floor Heights")
    ws.freeze_panes = "A2"
    _header_row(ws, ["Annotation", "Level (m)", "Layer", "Source File", "Full Description"],
                    [26, 12, 26, 22, 52])
    for i, fh in enumerate(items, 1):
        level_m = _mm_to_m(fh["level_mm"]) if fh["level_mm"] is not None else None
        _data_row(ws, i + 1, [
            (fh["annotation"],  _left()),
            (level_m,           _center()),
            (fh["layer"],       _left()),
            (fh["source"],      _left()),
            (fh["description"], _left()),
        ], i)


# ── Public API ────────────────────────────────────────────────────────────────

def parse_dxf_to_excel(files: list[tuple[str, bytes]]) -> BytesIO:
    """
    Parse one or more DXF files and return an Excel workbook as BytesIO.

    Args:
        files: list of (filename, raw_bytes) tuples.
    """
    all_dims:  list[dict] = []
    all_texts: list[dict] = []

    for filename, file_bytes in files:
        source = os.path.splitext(os.path.basename(filename))[0]
        try:
            doc = _load_doc(file_bytes)
        except Exception as exc:
            raise ValueError(f"Could not read DXF file '{filename}': {exc}") from exc

        scale = _get_scale(doc)
        texts = _collect_text_entities(doc, source)
        dims  = _collect_dimensions(doc, scale, source, texts)

        all_dims.extend(dims)
        all_texts.extend(texts)

    steel        = _extract_steel(all_texts)
    windows_doors = _extract_windows_doors(all_texts)
    floor_heights = _extract_floor_heights(all_texts)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _sheet_by_room(wb, all_dims)
    _sheet_all_dims(wb, all_dims)
    _sheet_steel(wb, steel)
    _sheet_windows_doors(wb, windows_doors)
    _sheet_floor_heights(wb, floor_heights)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
